from django.core.mail import EmailMessage
from django.http import HttpResponse
from django.core.exceptions import ValidationError
from django.conf import settings
import os
from openpyxl import load_workbook, Workbook
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from .models import Teacher

def process_pdf_and_send_emails(modeladmin, request, queryset):
    # Iterate through selected Excel files
    for excel_file in queryset:
        # Load the Excel file
        file_path = excel_file.file.path
        try:
            wb = load_workbook(file_path, data_only=True)
        except Exception as e:
            raise ValidationError(f"Error loading the Excel file: {str(e)}")

        # Iterate through sheets in the Excel file
        for sheet_name in wb.sheetnames:
            # Check if a teacher with this sheet_name exists
            teachers = Teacher.objects.filter(sheet_name=sheet_name)

            if not teachers:
                continue

            # Collect email addresses of all teachers for this sheet
            recipient_list = [teacher.email for teacher in teachers]

            # Create a PDF for the sheet
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            elements = []

            sheet = wb[sheet_name]
            data = []

            for row in sheet.iter_rows(values_only=True):
                data.append(row)

            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),  # Header row background color
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),  # Header text color
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Centered alignment
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),  # Table body background color
                ('GRID', (0, 0), (-1, -1), 1, colors.black)  # Gridlines
            ]))

            elements.append(table)
            doc.build(elements)

            # Send the PDF to the teachers' emails
            subject = f'Your PDF file'
            message = 'Please find the attached PDF file containing your Excel sheet.'
            from_email = os.environ.get("EMAIL_FROM")

            pdf = buffer.getvalue()
            buffer.close()

            email = EmailMessage(subject, message, from_email, recipient_list)
            email.attach(f'{sheet_name}.pdf', pdf, 'application/pdf')
            email.send()

    # Provide feedback to the admin user
    modeladmin.message_user(request, "PDFs generated and emails sent for selected Excel files.")

# Define a description for the custom admin action
process_pdf_and_send_emails.short_description = "Generate PDFs and Send Emails"



def process_sheet_and_send_emails(modeladmin, request, queryset):
    # Iterate through selected Excel files
    for excel_file in queryset:
        # Load the Excel file
        file_path = excel_file.file.path
        try:
            wb = load_workbook(file_path, data_only=True)
        except Exception as e:
            raise ValidationError(f"Error loading the Excel file: {str(e)}")

        # Iterate through sheets in the Excel file
        for sheet_name in wb.sheetnames:
            # Check if a teacher with this sheet_name exists
            teachers = Teacher.objects.filter(sheet_name = sheet_name)

            if not teachers:
                continue
            
            emails = [teacher.email for teacher in teachers]

            # Create a new Excel workbook containing only this sheet
            buffer = BytesIO()
            sheet = wb[sheet_name]
            new_wb = Workbook()
            new_ws = new_wb.active

            for row in sheet.iter_rows():
                for cell in row:
                    new_ws[cell.coordinate] = cell.value

            new_wb.save(buffer)
            
            # Iterate through each teacher and send the Excel sheet
            subject = f'Your Excel Sheet'
            message = 'Here is your Excel sheet.'
            from_email = os.environ.get("EMAIL_FROM")
            recipient_list = emails
            excel_sheet = buffer.getvalue()
            buffer.close()

            email = EmailMessage(subject, message, from_email, recipient_list)
            email.attach(f'{sheet_name}.xlsx', excel_sheet, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            email.send()


    # Provide feedback to the admin user
    modeladmin.message_user(request, "Emails sent for selected Excel files.")


# Define a description for the custom admin action
process_sheet_and_send_emails.short_description = "Generate Sheets and Send Emails"